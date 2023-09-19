# -*- encoding: utf-8 -*-
# @File     : __init__
# @Time     : 2023-09-16 01:19:31
# @Docs     : 数据提取解析
import json
from pathlib import Path
from functools import wraps

import xlrd
from icecream import ic
from loguru import logger
from openpyxl.workbook import Workbook
from PIL import ImageFont


__all__ = ("Config", "log", "MethodLogger")


# 常量
BASEDIR: Path = Path(__file__).resolve().parents[1]
RESDIR: Path = BASEDIR.joinpath("res")
LOGDIR: Path = BASEDIR.joinpath("logs")
GENDIR: Path = BASEDIR.joinpath("gen")
IMGDIR: Path = GENDIR.joinpath("img")
DATADIR: Path = GENDIR.joinpath("data")
FONTDIR: Path = RESDIR.joinpath("font")
TESTDIR: Path = RESDIR.joinpath("test")
CACHEDIR: Path = RESDIR.joinpath("cache")


# 日志
log = logger.bind(name="CnameApp")
log.add(
    LOGDIR.joinpath("spider.log"),
    format="{time} | {level} | {message}",
    rotation="1 week",
)


class Config:
    """
    配置类
    """

    def __init__(self, file: str = "config.json") -> None:
        self.conf_path = BASEDIR.joinpath(file)
        self._dynamic_create

    @property
    def _dynamic_create(self):
        """创建目录（如果不存在）

        Returns:
            _type_: bool
        """
        iter_dir = [GENDIR, RESDIR, LOGDIR, FONTDIR,
                    CACHEDIR, TESTDIR, IMGDIR, DATADIR]
        for dir in iter_dir:
            if not dir.exists():
                dir.mkdir()

    def _init_config(self):
        """初始化配置

        """
        tmpl = {
            "username": "",
            "password": "",
        }

        try:
            with open(self.conf_path, mode="w", encoding="utf-8") as file_obj:
                content: str = json.dumps(tmpl, ensure_ascii=False)
                file_obj.write(content)
        except Exception as err:
            log.exception(err)

    def load_config(self) -> dict:
        """导入配置文件

        Returns:
            _type_: dict
        """
        try:
            if not self.conf_path.exists():
                self._init_config()
                raise ValueError("请填写配置")

            result = self.load_json_file(self.conf_path)
            if result.get("username") == "":
                raise ValueError("请填写配置")
            return result

        except (FileNotFoundError, ValueError) as err:
            log.exception(err)
        return {}

    def load_font(self, font_name: str = "Deng", font_size: int = 30) -> ImageFont.FreeTypeFont:
        """ 导入字体文件 """
        if not (font_path := FONTDIR.joinpath(f"{font_name}.ttf")).exists():
            raise FileNotFoundError(f"{font_name} 不存在")
        return ImageFont.truetype(f"{font_path}", size=font_size)

    @classmethod
    def save_json_file(cls, save_file, content):
        with open(save_file, mode="w", encoding="utf-8") as file_obj:
            file_obj.write(json.dumps(content, ensure_ascii=False))

    @classmethod
    def load_json_file(cls, file_path) -> dict:
        with open(file_path, mode="r", encoding="utf-8") as file_obj:
            return json.loads(file_obj.read())


class MethodLogger:
    def __init__(self, skip_log=True, skip_save=True):
        self.skip_log = skip_log
        self.skip_save = skip_save

    def __call__(self, func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            result = func(*args, **kwargs)

            if not self.skip_log:
                log.info(f"Method {func.__name__} :=> {result}")
            else:
                ic(result)

            if not self.skip_save:
                self.save_result(func.__name__, result)

            return result

        return wrapper

    def save_result(self, filename: str, data_info) -> Path:
        def dump(x): return json.dumps(x or "{}", ensure_ascii=False)
        suffix = ".json"

        if isinstance(data_info, dict):
            content = dump(data_info)
        elif isinstance(data_info, str):
            suffix = ".txt"
            content = data_info
        elif isinstance(data_info, list):
            content = {"funcName": filename}
            content.update({str(k): v for k, v in enumerate(data_info)})
            content = dump(content)
        else:
            content = data_info or ""

        save_file = CACHEDIR.joinpath(f"{filename}{suffix}")
        Config.save_json_file(
            save_file,
            content,
        )
        return save_file


class ViewInfo:
    TableHref = "body > div.wap > a:nth-child(8)"
    LoginPageURL = "https://webvpn.jsu.edu.cn/https/77726476706e69737468656265737421e0f6528f693a7b45300d8db9d6562d/#/UserLogin?sn=ELNmlR8NQXCY1q-kHEz7xA&client_id=lFkRfDkwSW6z3IDAZpZo3g&redirect_uri=https%3A%2F%2Fwebvpn.jsu.edu.cn%2Flogin%3Foauth_login%3Dtrue"
    IndexPageWait = "#app > div > div:nth-child(2) > div:nth-child(1) > div > div.w50.mr20.pdt10 > div > div > div.fsb.mt10 > div"
    IndexPeopleInfo = '//*[@id="app"]/div/div[2]/div[1]/div/div[1]/div/div/div[2]/div'
    IndexWeatherPage = '//*[@id="app"]/div/div[2]/div[1]/div/div[1]/div/div/div[2]/iframe'


class ErrorStatus:
    ParserError = "课表数据解析错误!"
    ServerError = "教务系统服务可能出错了!"
    NotFoundBaseIMG = "没有初始文件, 正在绘制。"


def open_xls_as_xlsx(xls_path, xlsx_path) -> Path:
    """Excel文件转换

    Args:
        xls_path (_type_): 源文件
        xlsx_path (_type_): 目标文件

    Returns:
        _type_: 目标文件路径
    """
    try:
        book = xlrd.open_workbook(xls_path)
        index = 0
        nrows, ncols = 0, 0
        sheet = book.sheet_by_index(0)
        while nrows * ncols == 0:
            sheet = book.sheet_by_index(index)
            nrows = sheet.nrows
            ncols = sheet.ncols
            index += 1

        # 准备一个xlsx sheet
        book_new = Workbook()
        sheet_new = book_new.create_sheet("sheet1", 0)

        for row in range(0, nrows):
            for col in range(0, ncols):
                sheet_new.cell(row=row + 1, column=col + 1).value = sheet.cell_value(
                    row, col
                )

        book_new.save(xlsx_path)
        return xlsx_path
    except Exception as err:
        log.exception(err)
        raise err
